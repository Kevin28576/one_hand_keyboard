import struct
import threading
import time
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from collections import deque
from datetime import datetime
from pathlib import Path
import os
import json

BACKEND = None
try:
    import hid # type: ignore

    BACKEND = "hidapi"
except Exception:
    hid = None

if BACKEND is None:
    try:
        import pywinusb.hid as win_hid # type: ignore

        BACKEND = "pywinusb"
    except Exception as exc:
        raise SystemExit(
            "HID 函式庫載入失敗，請先安裝相依套件：\n"
            "pip install -r tools/requirements.txt\n"
            "Python 3.13 會自動使用 pywinusb 作為備援後端。"
        ) from exc


REPORT_ID_GAMEPAD = 6
USAGE_PAGE_GENERIC_DESKTOP = 0x01
USAGE_JOYSTICK = 0x04
APP_DIR = Path(os.getenv("APPDATA", ".")) / "OneHandKeyboard"
SETTINGS_PATH = APP_DIR / "monitor_settings.json"
LOG_DIR = APP_DIR / "logs"

KEYMAPS = [
    [
        "空", "~", "[", "]", "-", "=", ";", "\\",
        "ESC", "6", "7", "8", "9", "0", ",", ".",
        "CAPS", "1", "2", "3", "4", "5", "/", "'",
        "TAB", "Y", "Q", "W", "E", "R", "T", "BACKSPACE",
        "LSHIFT", "H", "A", "S", "0x07", "F", "G", "ENTER",
        "空", "Z", "X", "C", "V", "LALT", "LCTRL", "LAYER",
        "LEFT", "RIGHT", "SPACE", "空", "空", "空", "空", "空",
    ],
    [
        "空", "空", "空", "空", "空", "空", "空", "空",
        "LWIN", "F7", "F8", "F9", "F10", "F11", "F12", "空",
        "CAPS", "F1", "F2", "F3", "F4", "F5", "F6", "空",
        "TAB", "空", "UP", "U", "I", "O", "P", "DELETE",
        "LSHIFT", "LEFT", "DOWN", "RIGHT", "J", "K", "L", "ENTER",
        "空", "空", "B", "N", "M", "空", "功能", "功能",
        "空", "空", "空", "空", "空", "空", "空", "空",
    ],
    [
        "空", "1", "S", "F", "Y", "8", "L", "/",
        "ESC", "Q", "X", "V", "H", "I", ".", "-",
        "CAPS", "A", "E", "5", "N", "K", "0", "~",
        "TAB", "Z", "D", "T", "U", ",", "P", "BACKSPACE",
        "LSHIFT", "2", "C", "G", "J", "9", ";", "ENTER",
        "空", "W", "R", "B", "M", "O", "LCTRL", "LAYER",
        "LEFT", "RIGHT", "SPACE", "空", "空", "空", "空", "空",
    ],
    [
        "空", "空", "功能", "功能", "空", "空", "空", "空",
        "LWIN", "功能", "功能", "功能", "=", "功能", "空", "空",
        "CAPS", "SPACE", "6", "3", "4", "7", "功能", "空",
        "TAB", "功能", "功能", "功能", "功能", "空", "空", "DELETE",
        "LSHIFT", "空", "空", "空", "空", "空", "空", "空",
        "空", "空", "空", "空", "空", "空", "功能", "功能",
        "LEFT", "RIGHT", "SPACE", "空", "空", "空", "空", "空",
    ],
]

LAYER_NAMES_ZH = {
    0: "英文層",
    1: "英文層（FN）",
    2: "注音層",
    3: "注音層（FN）",
}

SPECIAL_KEY_LABELS = {
    40: "FN",
    47: "LAYER",
    48: "滑鼠左鍵",
    49: "滑鼠右鍵",
}

LAYER3_SPECIAL_LABELS = {
    46: "WIN+SPACE",
    2: "CTRL+[",
    3: "CTRL+]",
    9: "SHIFT+3",
    10: "SHIFT+7",
    11: "SHIFT+8",
    13: "SHIFT+=",
    22: "SHIFT+6",
    25: "CTRL+;",
    26: "CTRL+'",
    27: "CTRL+,",
    28: "CTRL+.",
}

LAYER_SPECIAL_LABELS = {
    1: {46: "WIN+SPACE"},
    3: LAYER3_SPECIAL_LABELS,
}


def get_key_label(layer, key_id):
    if key_id in SPECIAL_KEY_LABELS:
        return SPECIAL_KEY_LABELS[key_id]
    if layer in LAYER_SPECIAL_LABELS and key_id in LAYER_SPECIAL_LABELS[layer]:
        return LAYER_SPECIAL_LABELS[layer][key_id]
    if 0 <= layer < len(KEYMAPS) and 0 <= key_id < len(KEYMAPS[layer]):
        return KEYMAPS[layer][key_id]
    return "未知"


def parse_report(data):
    if not data:
        return None

    if len(data) >= 16 and data[0] == REPORT_ID_GAMEPAD:
        payload = bytes(data[1:16])
    elif len(data) >= 15:
        payload = bytes(data[:15])
    else:
        return None

    buttons, x, y, rx, ry, z, rz, dpads = struct.unpack("<IhhhhbbB", payload)
    dpad1 = dpads & 0x0F
    dpad2 = (dpads >> 4) & 0x0F

    return {
        "buttons": buttons,
        "key_press_count": x if x >= 0 else 0,
        "fn_press_count": y if y >= 0 else 0,
        "encoder_turn_count": rx if rx >= 0 else 0,
        "mouse_click_count": ry if ry >= 0 else 0,
        "current_layer": z if z >= 0 else 0,
        "last_key_id": rz if rz >= 0 else 0,
        "last_key_layer": dpad1 - 1 if dpad1 > 0 else None,
        "current_layer_dpad": dpad2 - 1 if dpad2 > 0 else None,
    }


class PyWinUsbDevice:
    def __init__(self, device):
        self.device = device
        self.queue = []
        self.lock = threading.Lock()
        self.event = threading.Event()

    def open(self):
        self.device.open()
        self.device.set_raw_data_handler(self._handler)

    def close(self):
        try:
            self.device.set_raw_data_handler(None)
        except Exception:
            pass
        try:
            self.device.close()
        except Exception:
            pass

    def _handler(self, data):
        with self.lock:
            self.queue.append(data)
            self.event.set()

    def read(self, size, timeout_ms=200):
        if not self.event.wait(timeout_ms / 1000.0):
            return []
        with self.lock:
            data = self.queue.pop(0) if self.queue else []
            if not self.queue:
                self.event.clear()
            return data


def enumerate_gamepads():
    devices = []
    if BACKEND == "hidapi":
        for dev in hid.enumerate():
            if dev.get("usage_page") != USAGE_PAGE_GENERIC_DESKTOP:
                continue
            if dev.get("usage") != USAGE_JOYSTICK:
                continue
            dev["backend"] = "hidapi"
            devices.append(dev)
    else:
        for dev in win_hid.HidDeviceFilter(
            usage_page=USAGE_PAGE_GENERIC_DESKTOP, usage=USAGE_JOYSTICK
        ).get_devices():
            devices.append(
                {
                    "backend": "pywinusb",
                    "path": dev.device_path,
                    "vendor_id": dev.vendor_id,
                    "product_id": dev.product_id,
                    "product_string": dev.product_name,
                    "device": dev,
                }
            )
    return devices


class HIDLayerMonitorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("單手鍵盤 狀態監控")

        self.devices = []
        self.device = None
        self.reader_thread = None
        self.reader_running = False
        self.latest_data = None
        self.last_layer = None
        self.toast_window = None
        self.recent_keys = deque(maxlen=10)
        self.last_key_press_count = None
        self.last_press_time = None
        self.press_history = deque()
        self.stats_history = deque()
        self.logging_enabled = False
        self.csv_file = None
        self.json_file = None
        self.keymap_window = None
        self.heatmap_window = None
        self.heatmap_cells = []
        self.heatmap_canvas = None
        self.topmost_enabled = False
        self.auto_connect_enabled = True
        self.tray_enabled = False
        self.tray_icon = None
        self.per_key_counts = [0] * 56
        self.per_layer_counts = [[0] * 56 for _ in range(4)]
        self.session_rows = []

        self._build_ui()
        self.refresh_devices()
        self.root.after(100, self.update_ui)
        self._load_settings()
        self._auto_connect_if_possible()

    def _build_ui(self):
        # 簡潔風格，字體清楚好讀
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Title.TLabel", font=("Microsoft JhengHei", 16, "bold"))
        style.configure("Value.TLabel", font=("Microsoft JhengHei", 12))
        style.configure("Label.TLabel", font=("Microsoft JhengHei", 11))
        style.configure("Primary.TButton", font=("Microsoft JhengHei", 10, "bold"))

        frame = ttk.Frame(self.root, padding=12)
        frame.grid(row=0, column=0, sticky="nsew")

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        ttk.Label(frame, text="單手鍵盤 狀態監控", style="Title.TLabel").grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 8)
        )

        ttk.Label(frame, text="裝置：", style="Label.TLabel").grid(
            row=1, column=0, sticky="w"
        )

        self.device_combo = ttk.Combobox(frame, state="readonly", width=52)
        self.device_combo.grid(row=1, column=1, sticky="we", padx=6)

        ttk.Button(frame, text="重新掃描", command=self.refresh_devices).grid(
            row=1, column=2, padx=6
        )
        self.connect_button = ttk.Button(
            frame, text="連線", command=self.toggle_connection, style="Primary.TButton"
        )
        self.connect_button.grid(row=1, column=3, padx=6)

        ttk.Separator(frame).grid(row=2, column=0, columnspan=4, sticky="we", pady=8)

        self.labels = {}
        row = 3
        for label in [
            "目前層（Layer）",
            "按鍵次數",
            "FN 使用次數",
            "旋鈕轉動次數",
            "滑鼠點擊次數",
            "最近按鍵 ID",
            "最近按鍵字符",
            "最近輸出內容",
            "最近按鍵座標",
            "最近按鍵所在層",
            "平均按鍵速度",
        ]:
            ttk.Label(frame, text=label + "：", style="Label.TLabel").grid(
                row=row, column=0, sticky="w"
            )
            value = ttk.Label(frame, text="-", style="Value.TLabel")
            value.grid(row=row, column=1, columnspan=3, sticky="w")
            self.labels[label] = value
            row += 1

        ttk.Label(frame, text="最近 10 次按鍵：", style="Label.TLabel").grid(
            row=row, column=0, sticky="w", pady=(6, 0)
        )
        self.recent_list = tk.Listbox(frame, height=5)
        self.recent_list.grid(row=row, column=1, columnspan=3, sticky="we", pady=(6, 0))
        row += 1

        action_frame = ttk.Frame(frame)
        action_frame.grid(row=row, column=0, columnspan=4, sticky="we", pady=(6, 0))
        self.topmost_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            action_frame, text="視窗置頂", variable=self.topmost_var, command=self._toggle_topmost
        ).grid(row=0, column=0, padx=(0, 10))
        self.autoconn_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            action_frame, text="自動連線", variable=self.autoconn_var, command=self._toggle_autoconn
        ).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(action_frame, text="鍵位表", command=self._open_keymap_window).grid(
            row=0, column=2, padx=(0, 10)
        )
        ttk.Button(action_frame, text="熱度圖", command=self._open_heatmap_window).grid(
            row=0, column=3, padx=(0, 10)
        )
        self.log_button = ttk.Button(
            action_frame, text="開始記錄", command=self._toggle_logging
        )
        self.log_button.grid(row=0, column=4, padx=(0, 10))
        ttk.Button(
            action_frame, text="匯出 Excel", command=self._export_excel
        ).grid(row=0, column=5, padx=(0, 10))
        ttk.Button(
            action_frame, text="匯出熱度PNG", command=self._export_heatmap_png
        ).grid(row=0, column=6, padx=(0, 10))
        ttk.Button(
            action_frame, text="最小化到托盤", command=self._minimize_to_tray
        ).grid(row=0, column=7)
        exit_btn = tk.Button(
            action_frame,
            text="退出",
            command=self._exit_app,
            font=("Microsoft JhengHei", 11, "bold"),
            bg="#d32f2f",
            fg="white",
            activebackground="#b71c1c",
            activeforeground="white",
            relief="raised",
            padx=14,
            pady=6,
        )
        exit_btn.grid(row=0, column=8, padx=(10, 0))

        self.status = ttk.Label(frame, text="狀態：未連線", style="Label.TLabel")
        self.status.grid(row=row + 1, column=0, columnspan=4, sticky="w", pady=(8, 0))

        frame.columnconfigure(1, weight=1)

    def _load_settings(self):
        if not SETTINGS_PATH.exists():
            return
        try:
            data = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
            self.autoconn_var.set(bool(data.get("auto_connect", True)))
            self.topmost_var.set(bool(data.get("topmost", False)))
            self._toggle_topmost()
        except Exception:
            return

    def _save_settings(self, selected_device=None):
        data = {
            "auto_connect": bool(self.autoconn_var.get()),
            "topmost": bool(self.topmost_var.get()),
        }
        if selected_device:
            data.update(selected_device)
        SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
        SETTINGS_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def _auto_connect_if_possible(self):
        if not self.autoconn_var.get():
            return
        if not SETTINGS_PATH.exists():
            return
        try:
            data = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
        except Exception:
            return
        vid = data.get("vendor_id")
        pid = data.get("product_id")
        if vid is None or pid is None:
            return
        for idx, dev in enumerate(self.devices):
            if dev.get("vendor_id") == vid and dev.get("product_id") == pid:
                self.device_combo.current(idx)
                self.connect()
                return

    def _toggle_topmost(self):
        self.topmost_enabled = bool(self.topmost_var.get())
        self.root.attributes("-topmost", self.topmost_enabled)
        self._save_settings()

    def _toggle_autoconn(self):
        self.auto_connect_enabled = bool(self.autoconn_var.get())
        self._save_settings()

    def _open_keymap_window(self):
        if self.keymap_window and tk.Toplevel.winfo_exists(self.keymap_window):
            self.keymap_window.lift()
            return
        self.keymap_window = tk.Toplevel(self.root)
        self.keymap_window.title("鍵位表")
        notebook = ttk.Notebook(self.keymap_window)
        notebook.pack(fill="both", expand=True, padx=8, pady=8)

        for layer_index, layer in enumerate(KEYMAPS):
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=f"Layer {layer_index}")
            for r in range(7):
                for c in range(8):
                    idx = r * 8 + c
                    text = layer[idx]
                    label = ttk.Label(frame, text=text, width=10, anchor="center")
                    label.grid(row=r, column=c, padx=2, pady=2, sticky="nsew")
            for c in range(8):
                frame.columnconfigure(c, weight=1)

    def _open_heatmap_window(self):
        if self.heatmap_window and tk.Toplevel.winfo_exists(self.heatmap_window):
            self.heatmap_window.lift()
            return
        self.heatmap_window = tk.Toplevel(self.root)
        self.heatmap_window.title("按鍵熱度圖（每層矩陣熱區）")
        container = ttk.Frame(self.heatmap_window, padding=10)
        container.pack(fill="both", expand=True)
        info = ttk.Label(
            container, text="顏色越亮代表按鍵被使用次數越多，可切換不同層。"
        )
        info.pack(anchor="w")

        self.heatmap_layer_var = tk.IntVar(value=0)
        layer_select = ttk.Frame(container)
        layer_select.pack(anchor="w", pady=(6, 0))
        ttk.Label(layer_select, text="層級：").pack(side="left")
        for i in range(4):
            ttk.Radiobutton(
                layer_select,
                text=f"Layer {i}",
                variable=self.heatmap_layer_var,
                value=i,
                command=self._refresh_heatmap,
            ).pack(side="left", padx=4)

        canvas = tk.Canvas(container, width=520, height=360, bg="#1e1e1e")
        canvas.pack(fill="both", expand=True, pady=(6, 0))
        self.heatmap_canvas = canvas
        self.heatmap_cells = []
        cell_w = 60
        cell_h = 40
        for r in range(7):
            row_cells = []
            for c in range(8):
                x0 = 10 + c * (cell_w + 4)
                y0 = 10 + r * (cell_h + 4)
                x1 = x0 + cell_w
                y1 = y0 + cell_h
                rect = canvas.create_rectangle(
                    x0, y0, x1, y1, fill="#2c2c2c", outline="#3a3a3a"
                )
                text = canvas.create_text(
                    (x0 + x1) // 2,
                    (y0 + y1) // 2,
                    text=f"{r},{c}",
                    fill="#cfcfcf",
                    font=("Microsoft JhengHei", 9),
                )
                row_cells.append((rect, text))
            self.heatmap_cells.append(row_cells)
        self._refresh_heatmap()

    def _refresh_heatmap(self):
        if not self.heatmap_cells:
            return
        layer_index = 0
        if hasattr(self, "heatmap_layer_var"):
            layer_index = int(self.heatmap_layer_var.get())
        per_layer_counts = self.per_key_counts if layer_index == 0 else None
        if hasattr(self, "per_layer_counts"):
            per_layer_counts = self.per_layer_counts[layer_index]
        if per_layer_counts is None:
            per_layer_counts = self.per_key_counts
        max_count = max(per_layer_counts) if per_layer_counts else 1
        max_count = max(max_count, 1)
        for r in range(7):
            for c in range(8):
                key_id = r * 8 + c
                count = per_layer_counts[key_id]
                intensity = int(50 + 205 * (count / max_count))
                color = f"#{intensity:02x}{intensity:02x}ff"
                rect, text = self.heatmap_cells[r][c]
                self.heatmap_canvas.itemconfig(rect, fill=color)
                self.heatmap_canvas.itemconfig(text, text=f"{key_id}\n{count}")

    def _export_excel(self):
        try:
            import openpyxl # type: ignore
        except Exception:
            self.status.config(text="狀態：匯出 Excel 需安裝 openpyxl")
            return
        if not self.session_rows:
            self.status.config(text="狀態：沒有可匯出的資料")
            return
        file_path = filedialog.asksaveasfilename(
            title="匯出 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 檔案", "*.xlsx")],
        )
        if not file_path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "即時資料"
        ws.append(
            [
                "時間",
                "層級",
                "按鍵次數",
                "FN 次數",
                "旋鈕次數",
                "滑鼠點擊",
                "最近按鍵ID",
                "最近按鍵字符",
            ]
        )
        for row in self.session_rows:
            ws.append(row)

        ws2 = wb.create_sheet("按鍵統計")
        ws2.append(["按鍵ID", "使用次數"])
        for key_id, count in enumerate(self.per_key_counts):
            ws2.append([key_id, count])

        ws3 = wb.create_sheet("鍵位對照表")
        ws3.append(["層級", "KeyID", "Row", "Col", "顯示文字"])
        for layer_index, layer in enumerate(KEYMAPS):
            for key_id, label in enumerate(layer):
                row = key_id // 8
                col = key_id % 8
                ws3.append([layer_index, key_id, row, col, label])

        ws4 = wb.create_sheet("每層熱度統計")
        ws4.append(["層級", "KeyID", "使用次數"])
        for layer_index, counts in enumerate(self.per_layer_counts):
            for key_id, count in enumerate(counts):
                ws4.append([layer_index, key_id, count])

        wb.save(file_path)
        self.status.config(text=f"狀態：已匯出 {Path(file_path).name}")

    def _export_heatmap_png(self):
        try:
            from PIL import Image, ImageDraw, ImageFont
        except Exception:
            self.status.config(text="狀態：匯出 PNG 需安裝 Pillow")
            return
        folder = filedialog.askdirectory(title="選擇輸出資料夾")
        if not folder:
            return

        cell_w = 70
        cell_h = 48
        margin = 12
        width = margin * 2 + 8 * cell_w + 7 * 4
        height = margin * 2 + 7 * cell_h + 6 * 4 + 50
        font = ImageFont.load_default()
        legend_x = margin
        legend_y = height - 28

        for layer_index, counts in enumerate(self.per_layer_counts):
            max_count = max(counts) if counts else 1
            max_count = max(max_count, 1)
            img = Image.new("RGB", (width, height), color="#1e1e1e")
            draw = ImageDraw.Draw(img)
            draw.text((margin, 6), f"Layer {layer_index} 熱度圖", fill="#ffffff", font=font)

            for r in range(7):
                for c in range(8):
                    key_id = r * 8 + c
                    count = counts[key_id]
                    intensity = int(50 + 205 * (count / max_count))
                    color = (intensity, intensity, 255)
                    x0 = margin + c * (cell_w + 4)
                    y0 = margin + 20 + r * (cell_h + 4)
                    x1 = x0 + cell_w
                    y1 = y0 + cell_h
                    draw.rectangle([x0, y0, x1, y1], fill=color, outline="#3a3a3a")
                    key_label = KEYMAPS[layer_index][key_id]
                    text = f"{key_id}\n{key_label}\n{count}"
                    draw.multiline_text((x0 + 3, y0 + 3), text, fill="#000000", font=font)

            # 色階圖例
            for i in range(10):
                ratio = i / 9
                intensity = int(50 + 205 * ratio)
                color = (intensity, intensity, 255)
                lx0 = legend_x + i * 18
                ly0 = legend_y
                lx1 = lx0 + 16
                ly1 = ly0 + 12
                draw.rectangle([lx0, ly0, lx1, ly1], fill=color, outline="#2a2a2a")
            draw.text((legend_x + 190, legend_y - 2), "低", fill="#ffffff", font=font)
            draw.text((legend_x + 230, legend_y - 2), "高", fill="#ffffff", font=font)
            draw.text((legend_x + 260, legend_y - 2), "使用次數", fill="#ffffff", font=font)
            draw.text((legend_x, legend_y + 14), "0", fill="#ffffff", font=font)
            draw.text(
                (legend_x + 160, legend_y + 14),
                str(max_count),
                fill="#ffffff",
                font=font,
            )

            out_path = Path(folder) / f"heatmap_layer_{layer_index}.png"
            img.save(out_path)

        self.status.config(text=f"狀態：已輸出熱度圖 PNG 到 {Path(folder).name}")

    def _toggle_logging(self):
        if self.logging_enabled:
            self._stop_logging()
        else:
            self._start_logging()

    def _start_logging(self):
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = LOG_DIR / f"telemetry_{timestamp}.csv"
        json_path = LOG_DIR / f"telemetry_{timestamp}.jsonl"
        self.csv_file = open(csv_path, "a", encoding="utf-8", newline="")
        self.json_file = open(json_path, "a", encoding="utf-8")
        self.csv_file.write(
            "time,layer,key_press_count,fn_press_count,encoder_turn_count,mouse_click_count,last_key_id,last_key_label\n"
        )
        self.logging_enabled = True
        self.log_button.config(text="停止記錄")
        self.status.config(text=f"狀態：記錄中（{csv_path.name}）")

    def _stop_logging(self):
        if self.csv_file:
            self.csv_file.close()
        if self.json_file:
            self.json_file.close()
        self.csv_file = None
        self.json_file = None
        self.logging_enabled = False
        self.log_button.config(text="開始記錄")

    def _minimize_to_tray(self):
        try:
            import pystray # type: ignore
            from PIL import Image, ImageDraw
        except Exception:
            self.status.config(text="狀態：托盤功能需要 pystray 與 Pillow")
            return

        if self.tray_icon:
            return

        def _create_image():
            image = Image.new("RGB", (64, 64), color="#2f2f2f")
            draw = ImageDraw.Draw(image)
            draw.rectangle((8, 8, 56, 56), outline="#ffffff", width=3)
            draw.text((18, 18), "KB", fill="#ffffff")
            return image

        def on_restore(icon, item):
            self.root.deiconify()
            self.root.attributes("-topmost", self.topmost_enabled)
            icon.stop()
            self.tray_icon = None

        def on_quit(icon, item):
            icon.stop()
            self.tray_icon = None
            self.disconnect()
            self.root.destroy()

        menu = pystray.Menu(
            pystray.MenuItem("還原", on_restore),
            pystray.MenuItem("結束", on_quit),
        )
        self.tray_icon = pystray.Icon("one_hand_keyboard", _create_image(), "單手鍵盤", menu)
        self.root.withdraw()
        threading.Thread(target=self.tray_icon.run, daemon=True).start()

    def refresh_devices(self):
        self.devices = enumerate_gamepads()
        values = []
        for idx, dev in enumerate(self.devices):
            product = dev.get("product_string") or "未知裝置"
            vid = dev.get("vendor_id")
            pid = dev.get("product_id")
            values.append(f"{idx}: {product} (VID {vid:04X}, PID {pid:04X})")
        self.device_combo["values"] = values
        if values:
            self.device_combo.current(0)
        self._auto_connect_if_possible()

    def toggle_connection(self):
        if self.device:
            self.disconnect()
        else:
            self.connect()

    def connect(self):
        if not self.devices:
            self.status.config(text="狀態：找不到裝置")
            return
        index = self.device_combo.current()
        if index < 0 or index >= len(self.devices):
            self.status.config(text="狀態：裝置選擇無效")
            return

        dev_info = self.devices[index]
        try:
            if dev_info.get("backend") == "pywinusb":
                self.device = PyWinUsbDevice(dev_info["device"])
                self.device.open()
            else:
                self.device = hid.Device(path=dev_info["path"])
            self.reader_running = True
            self.reader_thread = threading.Thread(
                target=self.reader_loop, daemon=True
            )
            self.reader_thread.start()
            self.connect_button.config(text="中斷")
            self.status.config(text="狀態：已連線")
            self._save_settings(
                {
                    "vendor_id": dev_info.get("vendor_id"),
                    "product_id": dev_info.get("product_id"),
                    "product_string": dev_info.get("product_string"),
                }
            )
        except Exception as exc:
            self.device = None
            self.status.config(text=f"狀態：連線失敗（{exc}）")

    def disconnect(self):
        self.reader_running = False
        if self.reader_thread:
            self.reader_thread.join(timeout=1.0)
        if self.device:
            try:
                self.device.close()
            except Exception:
                pass
        self.device = None
        self.reader_thread = None
        self.connect_button.config(text="連線")
        self.status.config(text="狀態：未連線")

    def reader_loop(self):
        while self.reader_running and self.device:
            try:
                data = self.device.read(64, timeout_ms=200)
                parsed = parse_report(data)
                if parsed:
                    self.latest_data = parsed
            except Exception:
                self.latest_data = None
                time.sleep(0.2)

    def _show_layer_toast(self, layer):
        # 顯示不干擾焦點的浮動提示
        if self.toast_window is not None:
            try:
                self.toast_window.destroy()
            except Exception:
                pass
            self.toast_window = None

        toast = tk.Toplevel(self.root)
        toast.overrideredirect(True)
        toast.attributes("-topmost", True)
        toast.attributes("-alpha", 0.9)
        try:
            toast.attributes("-toolwindow", True)
        except Exception:
            pass

        layer_name = LAYER_NAMES_ZH.get(layer, f"{layer}")

        label = tk.Label(
            toast,
            text=f"目前層：{layer_name}",
            font=("Microsoft JhengHei", 14, "bold"),
            bg="#1f1f1f",
            fg="#f0f0f0",
            padx=14,
            pady=10,
        )
        label.pack()

        toast.update_idletasks()
        screen_w = toast.winfo_screenwidth()
        screen_h = toast.winfo_screenheight()
        win_w = toast.winfo_width()
        win_h = toast.winfo_height()
        x = screen_w - win_w - 24
        y = screen_h - win_h - 72
        toast.geometry(f"{win_w}x{win_h}+{x}+{y}")

        self.toast_window = toast
        toast.after(1200, toast.destroy)

    def update_ui(self):
        data = self.latest_data
        if data:
            current_layer = data["current_layer"]
            last_key_id = data["last_key_id"]
            last_key_layer = data["last_key_layer"]
            row = last_key_id // 8
            col = last_key_id % 8
            key_layer = (
                last_key_layer if last_key_layer is not None else current_layer
            )
            key_label = get_key_label(key_layer, last_key_id)

            self.labels["目前層（Layer）"].config(text=str(current_layer))
            self.labels["按鍵次數"].config(text=str(data["key_press_count"]))
            self.labels["FN 使用次數"].config(text=str(data["fn_press_count"]))
            self.labels["旋鈕轉動次數"].config(
                text=str(data["encoder_turn_count"])
            )
            self.labels["滑鼠點擊次數"].config(text=str(data["mouse_click_count"]))
            self.labels["最近按鍵 ID"].config(text=str(last_key_id))
            self.labels["最近按鍵字符"].config(text=str(key_label))
            self.labels["最近輸出內容"].config(text=str(key_label))
            self.labels["最近按鍵座標"].config(text=f"R{row}C{col}")
            self.labels["最近按鍵所在層"].config(
                text=str(last_key_layer) if last_key_layer is not None else "-"
            )

            now = time.time()
            if self.last_key_press_count is None:
                self.last_key_press_count = data["key_press_count"]
                self.last_press_time = now
            if data["key_press_count"] != self.last_key_press_count:
                self.press_history.append((now, data["key_press_count"]))
                self.last_key_press_count = data["key_press_count"]
                self.last_press_time = now
                self.recent_keys.appendleft(f"{key_label} (L{key_layer})")
                self.recent_list.delete(0, tk.END)
                for item in self.recent_keys:
                    self.recent_list.insert(tk.END, item)
                if 0 <= last_key_id < len(self.per_key_counts):
                    self.per_key_counts[last_key_id] += 1
                    if 0 <= key_layer < len(self.per_layer_counts):
                        self.per_layer_counts[key_layer][last_key_id] += 1
                if self.heatmap_window and tk.Toplevel.winfo_exists(self.heatmap_window):
                    self._refresh_heatmap()
                ts = datetime.now().isoformat(timespec="seconds")
                self.session_rows.append(
                    [
                        ts,
                        current_layer,
                        data["key_press_count"],
                        data["fn_press_count"],
                        data["encoder_turn_count"],
                        data["mouse_click_count"],
                        last_key_id,
                        key_label,
                    ]
                )
                if self.logging_enabled and self.csv_file and self.json_file:
                    self.csv_file.write(
                        f"{ts},{current_layer},{data['key_press_count']},{data['fn_press_count']},"
                        f"{data['encoder_turn_count']},{data['mouse_click_count']},"
                        f"{last_key_id},{key_label}\n"
                    )
                    self.csv_file.flush()
                    self.json_file.write(
                        json.dumps(
                            {
                                "time": ts,
                                "layer": current_layer,
                                "key_press_count": data["key_press_count"],
                                "fn_press_count": data["fn_press_count"],
                                "encoder_turn_count": data["encoder_turn_count"],
                                "mouse_click_count": data["mouse_click_count"],
                                "last_key_id": last_key_id,
                                "last_key_label": key_label,
                            },
                            ensure_ascii=False,
                        )
                        + "\n"
                    )
                    self.json_file.flush()

            while self.press_history and now - self.press_history[0][0] > 60:
                self.press_history.popleft()
            if len(self.press_history) >= 2:
                t0, c0 = self.press_history[0]
                t1, c1 = self.press_history[-1]
                rate = (c1 - c0) / max(t1 - t0, 0.1)
                self.labels["平均按鍵速度"].config(text=f"{rate:.2f} 次/秒")
            else:
                self.labels["平均按鍵速度"].config(text="0.00 次/秒")

            if self.last_layer is None or current_layer != self.last_layer:
                self._show_layer_toast(current_layer)
                self.last_layer = current_layer
        else:
            for label in self.labels.values():
                label.config(text="-")

        self.root.after(100, self.update_ui)
    def _exit_app(self):
        # 正式退出：先停記錄、斷線、再關窗
        try:
            if self.logging_enabled:
                self._stop_logging()
        except Exception:
            pass

        try:
            self.disconnect()
        except Exception:
            pass

        try:
            self.root.quit()
        except Exception:
            pass

        self.root.destroy()


def main():
    root = tk.Tk()
    app = HIDLayerMonitorApp(root)
    root.protocol("WM_DELETE_WINDOW", app._exit_app)
    root.mainloop()


if __name__ == "__main__":
    main()
